import os
import uuid
import tempfile
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formula.translate import Translator
from transformers import AutoTokenizer, AutoModelForCausalLM
import torch
import xlwings as xw
from fastapi import FastAPI, UploadFile, File, Request
from fastapi.responses import FileResponse, JSONResponse, HTMLResponse
from pydantic import BaseModel
from fastapi.middleware.cors import CORSMiddleware
import tkinter as tk
from tkinter import filedialog, simpledialog, messagebox
from fastapi import APIRouter

router = APIRouter()

# कुछ routes भी होने चाहिए जैसे:
@router.get("/some-path")
def example_route():
    return {"message": "Hello"}

# ⚙️ FastAPI App
app = FastAPI()
TEMP_STORAGE = {}

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# 🤖 Load model (phi-1_5)
print("🔄 Loading AI model (phi-1_5)...")
tokenizer = AutoTokenizer.from_pretrained("microsoft/phi-1_5")
model = AutoModelForCausalLM.from_pretrained("microsoft/phi-1_5")
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
model.to(device)
print("✅ Model Ready")

# 🎯 AI summary from Hinglish
def generate_task_summary(user_cmd: str) -> str:
    prompt = f"User command: {user_cmd}\nWhat should be done (explain in 1 line):"
    inputs = tokenizer(prompt, return_tensors="pt").to(device)
    output = model.generate(**inputs, max_length=80)
    return tokenizer.decode(output[0], skip_special_tokens=True).split(":")[-1].strip()

# 🧮 Excel formula generator
def generate_formula_with_phi(instruction: str) -> str:
    prompt = f"Generate only Excel formula for: {instruction}\nFormula:"
    inputs = tokenizer(prompt, return_tensors="pt").to(device)
    output = model.generate(**inputs, max_length=100)
    return tokenizer.decode(output[0], skip_special_tokens=True).split("Formula:")[-1].strip()

# 📄 Apply formula to worksheet
def apply_formula_all_rows(ws, formula: str, start_row: int, target_col: int, max_row: int):
    ws.cell(row=1, column=target_col, value="Result")
    for i in range(start_row, max_row + 1):
        translated = Translator(formula, origin="B2").translate_formula(f"{chr(65 + target_col - 2)}{i}")
        ws.cell(row=i, column=target_col, value=f"={translated}")

# 🧠 AI logic to apply
def apply_excel_logic_with_formula(df: pd.DataFrame, instruction: str, sheet_name="Processed", overwrite=False, original_path=None) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    total_rows = len(df)
    total_cols = len(df.columns)

    table = Table(displayName="ExcelData", ref=ws.dimensions)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False,
                           showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    ws.add_table(table)

    formula = generate_formula_with_phi(instruction)
    if "=" in formula or any(c in formula for c in "+-*/SUMsum"):
        apply_formula_all_rows(ws, formula, start_row=2, target_col=total_cols + 1, max_row=total_rows + 1)

    if "pivot" in instruction.lower():
        print("ℹ️ Pivot logic: Not yet implemented")

    if overwrite and original_path:
        wb.save(original_path)
        print("♻️ File overwritten successfully.")
        return None  # Overwrite mode doesn't return new workbook

    return wb
# 🧾 Detect open workbook via xlwings
def get_open_excel_path():
    try:
        app = xw.App(visible=False)
        for wb in xw.apps.active.books:
            if wb.name.endswith('.xlsx'):
                return wb.fullname
    except Exception:
        return None

# 🌐 Web Interface (FastAPI Routes)
@app.get("/", response_class=HTMLResponse)
async def main_page():
    return HTMLResponse("""
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <title>📊 Smart Excel Assistant</title>
    <style>
        body {
            background: #f4f6f8;
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            padding: 50px;
            display: flex;
            justify-content: center;
        }
        .card {
            background: #fff;
            padding: 40px;
            border-radius: 12px;
            box-shadow: 0 8px 20px rgba(0, 0, 0, 0.1);
            width: 100%;
            max-width: 600px;
        }
        h2 {
            color: #1e88e5;
            text-align: center;
            margin-bottom: 30px;
        }
        input[type="text"] {
            width: 100%;
            padding: 12px;
            margin: 12px 0;
            border: 1px solid #ccc;
            border-radius: 6px;
            font-size: 16px;
        }
        input[type="submit"] {
            background-color: #1e88e5;
            color: white;
            padding: 14px;
            border: none;
            border-radius: 6px;
            font-size: 16px;
            width: 100%;
            cursor: pointer;
            margin-top: 10px;
        }
        input[type="submit"]:hover {
            background-color: #1565c0;
        }
        .footer {
            margin-top: 30px;
            text-align: center;
            font-size: 14px;
            color: #888;
        }
    </style>
</head>
<body>
    <div class="card">
        <h2>📊 Smart Excel Assistant</h2>
        <form action="/options/" method="post">
            <input type="text" name="session_id" placeholder="Leave blank to use open Excel">
            <input type="text" name="user_instruction" placeholder="e.g. Sum of sales column">
            <input type="text" name="sheet_name" placeholder="Sheet name (optional)">
            <input type="submit" value="🚀 Process Sheet">
        </form>
        <div class="footer">Developed by AI | Powered by phi-1_5 🤖</div>
    </div>
</body>
</html>
    """)

@app.post("/upload/")
async def upload_excel(file: UploadFile = File(...)):
    contents = await file.read()
    temp_id = str(uuid.uuid4())
    temp_path = os.path.join(tempfile.gettempdir(), f"{temp_id}.xlsx")
    with open(temp_path, "wb") as f:
        f.write(contents)
    TEMP_STORAGE[temp_id] = temp_path
    return {"message": "✅ File uploaded", "session_id": temp_id}

class Options(BaseModel):
    session_id: str = ""
    user_instruction: str
    sheet_name: str = "Processed"
    overwrite: bool = False

@app.post("/options/")
async def process_with_options(options: Options):
    if options.session_id:
        file_path = TEMP_STORAGE.get(options.session_id, "")
    else:
        file_path = get_open_excel_path()

    if not file_path or not os.path.exists(file_path):
        return JSONResponse(status_code=400, content={"error": "No valid Excel file found"})

    df = pd.read_excel(file_path)

    if options.overwrite:
        apply_excel_logic_with_formula(df, options.user_instruction, options.sheet_name, overwrite=True, original_path=file_path)
        return JSONResponse(content={"message": "✅ File overwritten successfully", "path": file_path})

    wb = apply_excel_logic_with_formula(df, options.user_instruction, options.sheet_name)
    output_path = file_path.replace(".xlsx", "_output.xlsx")
    wb.save(output_path)
    return FileResponse(output_path, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename="smart_processed.xlsx")

# 🖥️ CLI Mode
def cli_process(file_path: str, instruction: str, sheet_name="Processed", overwrite=False):
    df = pd.read_excel(file_path)
    if overwrite:
        apply_excel_logic_with_formula(df, instruction, sheet_name, overwrite=True, original_path=file_path)
        print(f"♻️ File overwritten: {file_path}")
    else:
        wb = apply_excel_logic_with_formula(df, instruction, sheet_name)
        output_path = file_path.replace(".xlsx", "_output_cli.xlsx")
        wb.save(output_path)
        print(f"✅ Excel processed: {output_path}")

# 🖱️ GUI Mode
def start_gui():
    root = tk.Tk()
    root.title("Smart Excel AI")

    def run():
        file_path = filedialog.askopenfilename()
        if not file_path:
            file_path = get_open_excel_path()
        instruction = simpledialog.askstring("Instruction", "क्या करना है (जैसे: sum of Sales)?")
        sheet_name = simpledialog.askstring("Sheet Name", "Sheet का नाम:", initialvalue="Processed")
        try:
            cli_process(file_path, instruction, sheet_name)
            messagebox.showinfo("Done", "File processed successfully!")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    tk.Button(root, text="Select and Process Excel", command=run, height=2, width=30).pack(pady=20)
    root.mainloop()

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="🛠️ Excel AI CLI Tool")
    parser.add_argument("--file", required=False, help="Path to Excel file (or auto-detect)")
    parser.add_argument("--instruction", required=False, help="Instruction for formula or task")
    parser.add_argument("--sheet", default="Processed", help="Sheet name")
    parser.add_argument("--gui", action="store_true", help="Launch GUI")
    parser.add_argument("--overwrite", action="store_true", help="Overwrite the original Excel file")

    args = parser.parse_args()
    excel_path = args.file or get_open_excel_path()

    if args.gui:
        start_gui()
    elif excel_path and args.instruction:
        cli_process(excel_path, args.instruction, args.sheet, overwrite=args.overwrite)
    else:
        print("📎 Use FastAPI for web mode, --gui for GUI, or pass --file and --instruction for CLI mode")